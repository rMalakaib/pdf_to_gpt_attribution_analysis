# Import necessary library
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import csv

FUNDNAME = "C+C"

# Create a new Workbook and select the active Worksheet
wb = Workbook()
ws = wb.active
ws.title = f"{FUNDNAME}"

def Format_Main_Table():
    # Define the data and headers
    headers = ["Index", "Company","Sector","Structure","Date","Invested","Realized","NAV (write current one cell above)","Total Value","MOIC","% Cost","% NAV","Entry Own%","Current Own%", "Lead","Region","Stage","BoD","Implied Entry Val (M)","Implied Current Val (M)","Avg Entry","Year","Tokens","Ticker","   ","Price","<ENTER DATE> NAV","CURRENT NAV","△ NAV","% △ NAV"]

    data = []

    i = 0
    with open('na.csv', mode='r') as file:
            # Create a CSV reader object
            csv_reader = csv.reader(file)
            
            for row in csv_reader:
                i = i + 1
                data.append([i])

            data.append(["TOTAL"])
            
            

    # Append headers to the worksheet starting at column A, row 4
    header_row = 4
    header_col = 1  # Column A

    for col_num, header in enumerate(headers, start=header_col):
        ws.cell(row=header_row, column=col_num, value=header)

    # Append data to the worksheet starting at column B, row 5
    data_start_row = 5

    for row_num, row_data in enumerate(data, start=data_start_row):
        for col_num, cell_value in enumerate(row_data, start=header_col+1):
            ws.cell(row=row_num, column=col_num-1, value=cell_value)

    
    # Add formula to calculate Total Compensation in each row
    for row_num in range(data_start_row, data_start_row + len(data)-1):
        
        # formatting NAV
        CURRENT_NAV = f"AB{row_num}"
        DATE_NAV = f"AA{row_num}"
        ACTUAL_NAV = f"H{row_num}"
        ws[ACTUAL_NAV] = f"=IF($H$3=\"current\",{DATE_NAV},{CURRENT_NAV})"

        # formatting total value
        REALIZED = f"G{row_num}"
        ACTUAL_NAV = f"H{row_num}"
        TOTAL_VALUE = f"I{row_num}"
        ws[TOTAL_VALUE] = f"={REALIZED}+{ACTUAL_NAV}"

        # formatting MOIC
        INVESTED = f"F{row_num}"
        MOIC = f"J{row_num}"
        ws[MOIC] = f"=IFERROR({TOTAL_VALUE}/{INVESTED},\"n/a\")"

        # formatting % cost
        TOTAL_INVESTED = f"$F${data_start_row+len(data)-1}"
        PERCENT_COST = f"K{row_num}"
        ws[PERCENT_COST] = f"={INVESTED}/{TOTAL_INVESTED}"

        # formatting % NAV
        TOTAL_NAV = f"$H${data_start_row+len(data)-1}"
        PERCENT_NAV = f"L{row_num}"
        ws[PERCENT_NAV] = f"={ACTUAL_NAV}/{TOTAL_NAV}"

        # formatting Implied Entry Valuation % (M)
        ENTRY_OWN_PERCENT = f"M{row_num}"
        IMPLIED_ENTRY = f"S{row_num}"
        ws[IMPLIED_ENTRY] = f"=IFERROR({INVESTED}/{ENTRY_OWN_PERCENT}/1000000,\"n/a\")"

        # formatting Current Entry Valuation % (M)
        CURRENT_OWN_PERCENT = f"N{row_num}"
        IMPLIED_CURRENT_VALUE = f"T{row_num}"
        ws[IMPLIED_CURRENT_VALUE] = f"=IFERROR({ACTUAL_NAV}/{CURRENT_OWN_PERCENT}/1000000,\"n/a\")"

        # formatting year of investment
        DATE = f"E{row_num}"
        INV_YEAR = f"V{row_num}"
        ws[INV_YEAR] = f"=YEAR({DATE})"

        # formatting change in NAV
        CHANGE_NAV = f"AC{row_num}"
        ws[CHANGE_NAV] = f"={DATE_NAV}-{CURRENT_NAV}"

        # formatting percent change in NAV between DATE and CURRENT
        PERCENTAGE_DATE_CURRENT = f"AD{row_num}"
        ws[PERCENTAGE_DATE_CURRENT] = f"={CHANGE_NAV}/{DATE_NAV}-1"

        # formatting Avg Entry
        AVG_ENTRY = f"U{row_num}"
        ws[AVG_ENTRY] = f"=IFERROR({IMPLIED_CURRENT_VALUE}*{INVESTED}/{TOTAL_VALUE},\"n/a\")"

    #TOTALS ROW
    # formatting invested total
    INVESTING_TOTAL = f"F{data_start_row+len(data)-1}"
    INVESTING_TOTAL_ROWS = f"F{data_start_row}:F{data_start_row+len(data)-2}"
    ws[INVESTING_TOTAL] = f"=SUM({INVESTING_TOTAL_ROWS})"

    # formatting realized total
    REALIZED_TOTAL = f"G{data_start_row+len(data)-1}"
    REALIZED_TOTAL_ROWS = f"G{data_start_row}:G{data_start_row+len(data)-2}"
    ws[REALIZED_TOTAL] = f"=SUM({REALIZED_TOTAL_ROWS})"

    # formatting NAV total
    NAV_TOTAL = f"H{data_start_row+len(data)-1}"
    NAV_TOTAL_ROWS = f"H{data_start_row}:H{data_start_row+len(data)-2}"
    ws[NAV_TOTAL] = f"=SUM({NAV_TOTAL_ROWS})"

    # formatting SUM TOTAL VALUE 
    TOTAL_VALUE_TOTAL = f"I{data_start_row+len(data)-1}"
    TOTAL_VALUE_TOTAL_ROWS = f"I{data_start_row}:I{data_start_row+len(data)-2}"
    ws[TOTAL_VALUE_TOTAL] = f"=SUM({TOTAL_VALUE_TOTAL_ROWS})"

    # formatting MOIC
    MOIC_ = f"J{data_start_row+len(data)-1}"
    ws[MOIC_] = f"=IFERROR({TOTAL_VALUE_TOTAL}/{INVESTING_TOTAL},\"n/a\")"

    # formatting % cost
    PERCENT_COST_ = f"K{data_start_row+len(data)-1}"
    ws[PERCENT_COST_] = f"={INVESTING_TOTAL}/{INVESTING_TOTAL}"

    # formatting % NAV
    PERCENT_NAV_ = f"L{data_start_row+len(data)-1}"
    ws[PERCENT_NAV_] = f"={NAV_TOTAL}/{NAV_TOTAL}"

    # misc equation
    IMPLIED_ENTRY_ROWS = f"S{data_start_row}:S{data_start_row+len(data)-2}"
    MISC_EQ = f"S{data_start_row+len(data)-1}"
    ws[MISC_EQ] = f"=SUMPRODUCT({INVESTING_TOTAL_ROWS},{IMPLIED_ENTRY_ROWS})/{INVESTING_TOTAL}"

    return data_start_row+len(data)-2


GLOBAL_COUNT = Format_Main_Table()

WB2=f"{FUNDNAME} Pivot Tables"
ws2 = wb.create_sheet(WB2)
wb.active = ws2

GLOBAL_DATA = [
    [1],
    [2],
    [3],
    [4],
    [5],
    [6],
    [7],
    [8],
    [9],
    [10],
    ["TOTAL"]
]




INVESTED = f"\'{FUNDNAME}\'!$F$5:$F${GLOBAL_COUNT}"
REALIZED = f"\'{FUNDNAME}\'!$G$5:$G${GLOBAL_COUNT}"
NAV = f"\'{FUNDNAME}\'!$H$5:$H${GLOBAL_COUNT}"
TOTAL_VALUE = f"\'{FUNDNAME}\'!$I$5:$I${GLOBAL_COUNT}"


def year():
    headers = ["Index","Year", "Count", "Invested", "Realized", "NAV", "Total Value", "MOIC", "% Cost", "% NAV"]
     
    # Append headers to the worksheet starting at column A, row 4
    header_row = 3
    header_col = 1  # Column A

    for col_num, header in enumerate(headers, start=header_col):
        ws2.cell(row=header_row, column=col_num, value=header)

    # Append data to the worksheet starting at column B, row 5
    data_start_row = 4

    # CHANGE
    YEAR = f"\'{FUNDNAME}\'!$V$5:$V${GLOBAL_COUNT}"

    for row_num, row_data in enumerate(GLOBAL_DATA, start=data_start_row):
        for col_num, cell_value in enumerate(row_data, start=header_col+1):
            ws2.cell(row=row_num, column=col_num-1, value=cell_value)

    for row_num in range(data_start_row, data_start_row + len(GLOBAL_DATA)-1):
        # count if year
        CAT_YEAR = f"B{row_num}"
        COUNT = f"C{row_num}"
        ws2[COUNT] = f"=COUNTIF({YEAR},\'{WB2}\'!{CAT_YEAR})"

        # sum if year by investment
        INVESTED_ = f"D{row_num}"
        ws2[INVESTED_] = f"=SUMIF({YEAR},\'{WB2}\'!{CAT_YEAR},{INVESTED})"

        # sum if year by realized
        REALIZED_ = f"E{row_num}"
        ws2[REALIZED_] = f"=SUMIF({YEAR},\'{WB2}\'!{CAT_YEAR},{REALIZED})"       

        # sum if year by NAV
        NAV_ = f"F{row_num}"
        ws2[NAV_] = f"=SUMIF({YEAR},\'{WB2}\'!{CAT_YEAR},{NAV})"   

        # sum if year by NAV
        TOTAL_VALUE_ = f"G{row_num}"
        ws2[TOTAL_VALUE_] = f"=SUMIF({YEAR},\'{WB2}\'!{CAT_YEAR},{TOTAL_VALUE})"   

        # MOIC calculation
        MOIC = f"H{row_num}"
        ws2[MOIC] = f"={TOTAL_VALUE_}/{INVESTED_}"

        # %Cost
        TOTAL_INVESTED = f"$D${GLOBAL_COUNT}"
        PERCENT_COST = f"I{row_num}"
        ws2[PERCENT_COST] = f"={INVESTED_}/{TOTAL_INVESTED}"

        # %NAV
        TOTAL_NAV = f"$F${GLOBAL_COUNT}"
        PERCENT_NAV = f"J{row_num}"
        ws2[PERCENT_NAV] = f"={NAV_}/{TOTAL_NAV}"

    # counting
    COUNTING_TOTAL = f"C{data_start_row+len(GLOBAL_DATA)-1}"
    COUNTING_SUM_ROWS = f"C{data_start_row}:C{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[COUNTING_TOTAL] = f"=SUM({COUNTING_SUM_ROWS})"
    
    # investing total rows
    INVESTING_TOTAL = f"D{data_start_row+len(GLOBAL_DATA)-1}"
    INVESTING_TOTAL_ROWS = f"D{data_start_row}:D{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[INVESTING_TOTAL] = f"=SUM({INVESTING_TOTAL_ROWS})"
     
    # formatting realized total
    REALIZED_TOTAL = f"E{data_start_row+len(GLOBAL_DATA)-1}"
    REALIZED_TOTAL_ROWS = f"E{data_start_row}:E{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[REALIZED_TOTAL] = f"=SUM({REALIZED_TOTAL_ROWS})"

    # formatting NAV total
    NAV_TOTAL = f"F{data_start_row+len(GLOBAL_DATA)-1}"
    NAV_TOTAL_ROWS = f"F{data_start_row}:F{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[NAV_TOTAL] = f"=SUM({NAV_TOTAL_ROWS})"

    # formatting SUM TOTAL VALUE 
    TOTAL_VALUE_TOTAL = f"G{data_start_row+len(GLOBAL_DATA)-1}"
    TOTAL_VALUE_TOTAL_ROWS = f"G{data_start_row}:G{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[TOTAL_VALUE_TOTAL] = f"=SUM({TOTAL_VALUE_TOTAL_ROWS})"

    # formatting MOIC
    MOIC_ = f"H{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[MOIC_] = f"=IFERROR({TOTAL_VALUE_TOTAL}/{INVESTING_TOTAL},\"n/a\")"

    # formatting % cost
    PERCENT_COST_ = f"I{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[PERCENT_COST_] = f"={INVESTING_TOTAL}/{INVESTING_TOTAL}"

    # formatting % NAV
    PERCENT_NAV_ = f"J{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[PERCENT_NAV_] = f"={NAV_TOTAL}/{NAV_TOTAL}"





def Stage():
    
    headers = ["Index","Stage", "Count", "Invested", "Realized", "NAV", "Total Value", "MOIC", "% Cost", "% NAV"]
    
    # Append headers to the worksheet starting at column A, row 4
    header_row = 17
    header_col = 1  # Column A

    for col_num, header in enumerate(headers, start=header_col):
        ws2.cell(row=header_row, column=col_num, value=header)

    # Append data to the worksheet starting at column B, row 5
    data_start_row = 18

    # CHANGE
    CAT = f"\'{FUNDNAME}\'!$Q${data_start_row}:$Q${GLOBAL_COUNT}"

    for row_num, row_data in enumerate(GLOBAL_DATA, start=data_start_row):
        for col_num, cell_value in enumerate(row_data, start=header_col+1):
            ws2.cell(row=row_num, column=col_num-1, value=cell_value)

    for row_num in range(data_start_row, data_start_row + len(GLOBAL_DATA)-1):
        # count if year
        CAT_STAGE = f"B{row_num}"
        COUNT = f"C{row_num}"
        ws2[COUNT] = f"=COUNTIF({CAT},\'{WB2}\'!{CAT_STAGE})"

        # sum if year by investment
        INVESTED_ = f"D{row_num}"
        ws2[INVESTED_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{INVESTED})"

        # sum if year by realized
        REALIZED_ = f"E{row_num}"
        ws2[REALIZED_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{REALIZED})"       

        # sum if year by NAV
        NAV_ = f"F{row_num}"
        ws2[NAV_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{NAV})"   

        # sum if year by NAV
        TOTAL_VALUE_ = f"G{row_num}"
        ws2[TOTAL_VALUE_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{TOTAL_VALUE})"   

        # MOIC calculation
        MOIC = f"H{row_num}"
        ws2[MOIC] = f"={TOTAL_VALUE_}/{INVESTED_}"

        # %Cost
        TOTAL_INVESTED = f"$D${GLOBAL_COUNT}"
        PERCENT_COST = f"I{row_num}"
        ws2[PERCENT_COST] = f"={INVESTED_}/{TOTAL_INVESTED}"

        # %NAV
        TOTAL_NAV = f"$F${GLOBAL_COUNT}"
        PERCENT_NAV = f"J{row_num}"
        ws2[PERCENT_NAV] = f"={NAV_}/{TOTAL_NAV}"
    
    # counting
    COUNTING_TOTAL = f"C{data_start_row+len(GLOBAL_DATA)-1}"
    COUNTING_SUM_ROWS = f"C{data_start_row}:C{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[COUNTING_TOTAL] = f"=SUM({COUNTING_SUM_ROWS})"

    # investing total rows
    INVESTING_TOTAL = f"D{data_start_row+len(GLOBAL_DATA)-1}"
    INVESTING_TOTAL_ROWS = f"D{data_start_row}:D{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[INVESTING_TOTAL] = f"=SUM({INVESTING_TOTAL_ROWS})"
     
    # formatting realized total
    REALIZED_TOTAL = f"E{data_start_row+len(GLOBAL_DATA)-1}"
    REALIZED_TOTAL_ROWS = f"E{data_start_row}:E{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[REALIZED_TOTAL] = f"=SUM({REALIZED_TOTAL_ROWS})"

    # formatting NAV total
    NAV_TOTAL = f"F{data_start_row+len(GLOBAL_DATA)-1}"
    NAV_TOTAL_ROWS = f"F{data_start_row}:F{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[NAV_TOTAL] = f"=SUM({NAV_TOTAL_ROWS})"

    # formatting SUM TOTAL VALUE 
    TOTAL_VALUE_TOTAL = f"G{data_start_row+len(GLOBAL_DATA)-1}"
    TOTAL_VALUE_TOTAL_ROWS = f"G{data_start_row}:G{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[TOTAL_VALUE_TOTAL] = f"=SUM({TOTAL_VALUE_TOTAL_ROWS})"

    # formatting MOIC
    MOIC_ = f"H{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[MOIC_] = f"=IFERROR({TOTAL_VALUE_TOTAL}/{INVESTING_TOTAL},\"n/a\")"

    # formatting % cost
    PERCENT_COST_ = f"I{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[PERCENT_COST_] = f"={INVESTING_TOTAL}/{INVESTING_TOTAL}"

    # formatting % NAV
    PERCENT_NAV_ = f"J{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[PERCENT_NAV_] = f"={NAV_TOTAL}/{NAV_TOTAL}"




def Geography():
    
    headers = ["Index","Geography", "Count", "Invested", "Realized", "NAV", "Total Value", "MOIC", "% Cost", "% NAV"]
    
    # Append headers to the worksheet starting at column A, row 4
    header_row = 31
    header_col = 1  # Column A

    for col_num, header in enumerate(headers, start=header_col):
        ws2.cell(row=header_row, column=col_num, value=header)

    # Append data to the worksheet starting at column B, row 5
    data_start_row = 32

    # CHANGE
    CAT = f"\'{FUNDNAME}\'!$P${data_start_row}:$P${GLOBAL_COUNT}"

    for row_num, row_data in enumerate(GLOBAL_DATA, start=data_start_row):
        for col_num, cell_value in enumerate(row_data, start=header_col+1):
            ws2.cell(row=row_num, column=col_num-1, value=cell_value)

    for row_num in range(data_start_row, data_start_row + len(GLOBAL_DATA)-1):
        # count if year
        CAT_STAGE = f"B{row_num}"
        COUNT = f"C{row_num}"
        ws2[COUNT] = f"=COUNTIF({CAT},\'{WB2}\'!{CAT_STAGE})"

        # sum if year by investment
        INVESTED_ = f"D{row_num}"
        ws2[INVESTED_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{INVESTED})"

        # sum if year by realized
        REALIZED_ = f"E{row_num}"
        ws2[REALIZED_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{REALIZED})"       

        # sum if year by NAV
        NAV_ = f"F{row_num}"
        ws2[NAV_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{NAV})"   

        # sum if year by NAV
        TOTAL_VALUE_ = f"G{row_num}"
        ws2[TOTAL_VALUE_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{TOTAL_VALUE})"   

        # MOIC calculation
        MOIC = f"H{row_num}"
        ws2[MOIC] = f"={TOTAL_VALUE_}/{INVESTED_}"


        # %Cost
        TOTAL_INVESTED = f"$D${GLOBAL_COUNT}"
        PERCENT_COST = f"I{row_num}"
        ws2[PERCENT_COST] = f"={INVESTED_}/{TOTAL_INVESTED}"

        # %NAV
        TOTAL_NAV = f"$F${GLOBAL_COUNT}"
        PERCENT_NAV = f"J{row_num}"
        ws2[PERCENT_NAV] = f"={NAV_}/{TOTAL_NAV}"
    
    # counting
    COUNTING_TOTAL = f"C{data_start_row+len(GLOBAL_DATA)-1}"
    COUNTING_SUM_ROWS = f"C{data_start_row}:C{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[COUNTING_TOTAL] = f"=SUM({COUNTING_SUM_ROWS})"

    # investing total rows
    INVESTING_TOTAL = f"D{data_start_row+len(GLOBAL_DATA)-1}"
    INVESTING_TOTAL_ROWS = f"D{data_start_row}:D{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[INVESTING_TOTAL] = f"=SUM({INVESTING_TOTAL_ROWS})"
     
    # formatting realized total
    REALIZED_TOTAL = f"E{data_start_row+len(GLOBAL_DATA)-1}"
    REALIZED_TOTAL_ROWS = f"E{data_start_row}:E{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[REALIZED_TOTAL] = f"=SUM({REALIZED_TOTAL_ROWS})"

    # formatting NAV total
    NAV_TOTAL = f"F{data_start_row+len(GLOBAL_DATA)-1}"
    NAV_TOTAL_ROWS = f"F{data_start_row}:F{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[NAV_TOTAL] = f"=SUM({NAV_TOTAL_ROWS})"

    # formatting SUM TOTAL VALUE 
    TOTAL_VALUE_TOTAL = f"G{data_start_row+len(GLOBAL_DATA)-1}"
    TOTAL_VALUE_TOTAL_ROWS = f"G{data_start_row}:G{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[TOTAL_VALUE_TOTAL] = f"=SUM({TOTAL_VALUE_TOTAL_ROWS})"

    # formatting MOIC
    MOIC_ = f"H{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[MOIC_] = f"=IFERROR({TOTAL_VALUE_TOTAL}/{INVESTING_TOTAL},\"n/a\")"

    # formatting % cost
    PERCENT_COST_ = f"I{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[PERCENT_COST_] = f"={INVESTING_TOTAL}/{INVESTING_TOTAL}"

    # formatting % NAV
    PERCENT_NAV_ = f"J{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[PERCENT_NAV_] = f"={NAV_TOTAL}/{NAV_TOTAL}"



def Structure():
    
    headers = ["Index","Structure", "Count", "Invested", "Realized", "NAV", "Total Value", "MOIC", "% Cost", "% NAV"]
    
    # Append headers to the worksheet starting at column A, row 4
    header_row = 45
    header_col = 1  # Column A

    for col_num, header in enumerate(headers, start=header_col):
        ws2.cell(row=header_row, column=col_num, value=header)

    # Append data to the worksheet starting at column B, row 5
    data_start_row = 46

    # CHANGE
    CAT = f"\'{FUNDNAME}\'!$D${data_start_row}:$D${GLOBAL_COUNT}"

    for row_num, row_data in enumerate(GLOBAL_DATA, start=data_start_row):
        for col_num, cell_value in enumerate(row_data, start=header_col+1):
            ws2.cell(row=row_num, column=col_num-1, value=cell_value)

    for row_num in range(data_start_row, data_start_row + len(GLOBAL_DATA)-1):
        # count if year
        CAT_STAGE = f"B{row_num}"
        COUNT = f"C{row_num}"
        ws2[COUNT] = f"=COUNTIF({CAT},\'{WB2}\'!{CAT_STAGE})"

        # sum if year by investment
        INVESTED_ = f"D{row_num}"
        ws2[INVESTED_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{INVESTED})"

        # sum if year by realized
        REALIZED_ = f"E{row_num}"
        ws2[REALIZED_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{REALIZED})"       

        # sum if year by NAV
        NAV_ = f"F{row_num}"
        ws2[NAV_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{NAV})"   

        # sum if year by NAV
        TOTAL_VALUE_ = f"G{row_num}"
        ws2[TOTAL_VALUE_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{TOTAL_VALUE})"   

        # MOIC calculation
        MOIC = f"H{row_num}"
        ws2[MOIC] = f"={TOTAL_VALUE_}/{INVESTED_}"


        # %Cost
        TOTAL_INVESTED = f"$D${GLOBAL_COUNT}"
        PERCENT_COST = f"I{row_num}"
        ws2[PERCENT_COST] = f"={INVESTED_}/{TOTAL_INVESTED}"

        # %NAV
        TOTAL_NAV = f"$F${GLOBAL_COUNT}"
        PERCENT_NAV = f"J{row_num}"
        ws2[PERCENT_NAV] = f"={NAV_}/{TOTAL_NAV}"
    
    # counting
    COUNTING_TOTAL = f"C{data_start_row+len(GLOBAL_DATA)-1}"
    COUNTING_SUM_ROWS = f"C{data_start_row}:C{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[COUNTING_TOTAL] = f"=SUM({COUNTING_SUM_ROWS})"

    # investing total rows
    INVESTING_TOTAL = f"D{data_start_row+len(GLOBAL_DATA)-1}"
    INVESTING_TOTAL_ROWS = f"D{data_start_row}:D{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[INVESTING_TOTAL] = f"=SUM({INVESTING_TOTAL_ROWS})"
     
    # formatting realized total
    REALIZED_TOTAL = f"E{data_start_row+len(GLOBAL_DATA)-1}"
    REALIZED_TOTAL_ROWS = f"E{data_start_row}:E{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[REALIZED_TOTAL] = f"=SUM({REALIZED_TOTAL_ROWS})"

    # formatting NAV total
    NAV_TOTAL = f"F{data_start_row+len(GLOBAL_DATA)-1}"
    NAV_TOTAL_ROWS = f"F{data_start_row}:F{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[NAV_TOTAL] = f"=SUM({NAV_TOTAL_ROWS})"

    # formatting SUM TOTAL VALUE 
    TOTAL_VALUE_TOTAL = f"G{data_start_row+len(GLOBAL_DATA)-1}"
    TOTAL_VALUE_TOTAL_ROWS = f"G{data_start_row}:G{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[TOTAL_VALUE_TOTAL] = f"=SUM({TOTAL_VALUE_TOTAL_ROWS})"

    # formatting MOIC
    MOIC_ = f"H{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[MOIC_] = f"=IFERROR({TOTAL_VALUE_TOTAL}/{INVESTING_TOTAL},\"n/a\")"

    # formatting % cost
    PERCENT_COST_ = f"I{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[PERCENT_COST_] = f"={INVESTING_TOTAL}/{INVESTING_TOTAL}"

    # formatting % NAV
    PERCENT_NAV_ = f"J{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[PERCENT_NAV_] = f"={NAV_TOTAL}/{NAV_TOTAL}"




def EntryOwenership():
    
    headers = ["Index","Entry Owenership", "Count", "Invested", "Realized", "NAV", "Total Value", "MOIC", "% Cost", "% NAV"]
    
    # Append headers to the worksheet starting at column A, row 4
    header_row = 59
    header_col = 1  # Column A

    for col_num, header in enumerate(headers, start=header_col):
        ws2.cell(row=header_row, column=col_num, value=header)

    # Append data to the worksheet starting at column B, row 5
    data_start_row = 60

    # CHANGE
    CAT = f"\'{FUNDNAME}\'!$M${data_start_row}:$M${GLOBAL_COUNT}"

    for row_num, row_data in enumerate(GLOBAL_DATA, start=data_start_row):
        for col_num, cell_value in enumerate(row_data, start=header_col+1):
            ws2.cell(row=row_num, column=col_num-1, value=cell_value)



    for row_num in range(data_start_row, data_start_row + len(GLOBAL_DATA)-1):
        # count if year
        COUNT = f"C{row_num}"

        if row_num == 60:
            ws2[COUNT] = f"=COUNTIF({CAT},\"<=1%\")"
        elif row_num == 61:
            ws2[COUNT] = f"=COUNTIF({CAT},\"<=5%\")-COUNTIF({CAT},\"<=1%\")"
        elif row_num == 62: 
            ws2[COUNT] = f"=COUNTIF({CAT},\">5%\")"
        else:
            ws2[COUNT] = f"=COUNTIF({CAT},\"n/a\")" 

        
        CAT_STAGE = f"B{row_num}"

        # sum if year by investment
        INVESTED_ = f"D{row_num}"
        ws2[INVESTED_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{INVESTED})"

        # sum if year by realized
        REALIZED_ = f"E{row_num}"
        ws2[REALIZED_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{REALIZED})"       

        # sum if year by NAV
        NAV_ = f"F{row_num}"
        ws2[NAV_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{NAV})"   

        # sum if year by NAV
        TOTAL_VALUE_ = f"G{row_num}"
        ws2[TOTAL_VALUE_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{TOTAL_VALUE})"   

        # MOIC calculation
        MOIC = f"H{row_num}"
        ws2[MOIC] = f"={TOTAL_VALUE_}/{INVESTED_}"


        # %Cost
        TOTAL_INVESTED = f"$D${GLOBAL_COUNT}"
        PERCENT_COST = f"I{row_num}"
        ws2[PERCENT_COST] = f"={INVESTED_}/{TOTAL_INVESTED}"

        # %NAV
        TOTAL_NAV = f"$F${GLOBAL_COUNT}"
        PERCENT_NAV = f"J{row_num}"
        ws2[PERCENT_NAV] = f"={NAV_}/{TOTAL_NAV}"
    
    # counting
    COUNTING_TOTAL = f"C{data_start_row+len(GLOBAL_DATA)-1}"
    COUNTING_SUM_ROWS = f"C{data_start_row}:C{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[COUNTING_TOTAL] = f"=SUM({COUNTING_SUM_ROWS})"

    # investing total rows
    INVESTING_TOTAL = f"D{data_start_row+len(GLOBAL_DATA)-1}"
    INVESTING_TOTAL_ROWS = f"D{data_start_row}:D{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[INVESTING_TOTAL] = f"=SUM({INVESTING_TOTAL_ROWS})"
     
    # formatting realized total
    REALIZED_TOTAL = f"E{data_start_row+len(GLOBAL_DATA)-1}"
    REALIZED_TOTAL_ROWS = f"E{data_start_row}:E{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[REALIZED_TOTAL] = f"=SUM({REALIZED_TOTAL_ROWS})"

    # formatting NAV total
    NAV_TOTAL = f"F{data_start_row+len(GLOBAL_DATA)-1}"
    NAV_TOTAL_ROWS = f"F{data_start_row}:F{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[NAV_TOTAL] = f"=SUM({NAV_TOTAL_ROWS})"

    # formatting SUM TOTAL VALUE 
    TOTAL_VALUE_TOTAL = f"G{data_start_row+len(GLOBAL_DATA)-1}"
    TOTAL_VALUE_TOTAL_ROWS = f"G{data_start_row}:G{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[TOTAL_VALUE_TOTAL] = f"=SUM({TOTAL_VALUE_TOTAL_ROWS})"

    # formatting MOIC
    MOIC_ = f"H{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[MOIC_] = f"=IFERROR({TOTAL_VALUE_TOTAL}/{INVESTING_TOTAL},\"n/a\")"

    # formatting % cost
    PERCENT_COST_ = f"I{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[PERCENT_COST_] = f"={INVESTING_TOTAL}/{INVESTING_TOTAL}"

    # formatting % NAV
    PERCENT_NAV_ = f"J{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[PERCENT_NAV_] = f"={NAV_TOTAL}/{NAV_TOTAL}"



EntryOwenership()
year()
Stage()
Geography()
Structure()

# Save the workbook
wb.save("./PreFormattedExcelTemplate.xlsx")
