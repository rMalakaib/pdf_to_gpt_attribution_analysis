from collections import namedtuple
import re
import pdfplumber
import pandas as pd
import openai
import json
import pandas as pd
import csv
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from config import *


# Set up the OpenAI API key
openai.api_key = SECRET_KEY

# CATEGORIES = "COMPANY NAME, CATEGORY, ENTRY DATE, COST, TOTAL VALUE, GROSS MOIC, ENTRY ROUND, ENTRY VALUATIONS (in millions), Entry %, Continent, Structure (equity or token)"

# CATEGORIES = "Company,Sector,Date,Investment"
# CATEGORIES = "COMPANY NAME, CATEGORY, ENTRY DATE, COST, TOTAL VALUE, GROSS MOIC, ENTRY ROUND, ENTRY VALUATIONS (in millions)"
# CATEGORIES = "Name,Type of Asset, Category, Description, Pre-Seed investment ($), Pre-Seed Investment Date, Pre-Seed Valuation (Post), Seed Investment ($), Seed Investment Date, Seed Valuation (post),  Internal Round Investment ($), Internal Round Investment Date, Internal Round Valuation (post), Series A investment ($), Series A Investment Date, Series A Valuation (post), Series A+ investment ($), Series A+ Investment Date, Series A+ Valuation (post), Series B investment ($), Series B Investment Date, Series B Valuation, Vested Tokens, Unvested Tokens, Earned Tokens, Market Purchased Tokens, Total Tokens, OTC/Market Purchase USD, Total Capital Committed ($), Current Equity FDV ($), Current Token FDV ($), Current NAV ($), Equity Ownership (%), Token Ownership (%)*"
def wait_for_categories():
    
    user_input = input("Input Categories like this   Company,Sector,Date,Investment  , hit enter to continue: ")
    return user_input

CATEGORIES = wait_for_categories()
NUMBER_CATEGORIES = str(len(CATEGORIES.split(",")))

# STEP 1
Line = namedtuple('Line', 'Name')
line_re = re.compile(r'ld \d{2,}')

def numbify(num):
    return float(num.replace('$', '').replace(',', ''))

data = []

def wait_for_pdf():
    
    user_input = input("Input PDF Path like this (the pdf must be located in the same folder as this script)   ./example.pdf  , hit enter to continue: ")
    return user_input

PDF_PATH = wait_for_pdf()

with pdfplumber.open(f"{PDF_PATH}") as pdf:
    for i in range(len(pdf.pages)):
        page = pdf.pages[i]
        text = page.extract_text(x_tolerance=2, y_tolerance=2)

        for line in text.split('\n'):

            if len(line) < 5:
                continue
            else: 
                data.append(line)
df = pd.DataFrame(data)    
df.to_csv("na.csv", index=False)

#STEP 1.5

def wait_for_yes():
    while True:
        user_input = input("After you've cleaned the data in na.csv, Type 'yes' to continue: ").strip().lower()
        if user_input == "yes":
            break
        else:
            print("Please type 'yes' to continue.")

wait_for_yes()
print(CATEGORIES)
print(NUMBER_CATEGORIES)


# STEP 2

def format_data_into_table(data):
    
    with open('na.csv', mode='r') as file:
            # Create a CSV reader object
            csv_reader = csv.reader(file)
            responses = []

            i = 0
            # Prepare the prompt
            # prompt = "formulate the data into a list delimited by | based off these categories: {CATEGORIES}. DO NOT CHANGE THE CATEGORY NAMES. OUTPUT ONLY A SINGLE ROW. ONLY RETURN each row as a single line. \n DO NOT USE backslash n. DO NOT WRITE TEXT ABOUT THE OUTPUT. \n DO NOT return single standalone letters that DO NOT fit into the respective categories, an example of this would be: [John Doe R 3.5x] TURNS INTO [John Doe | 3.5x].\n DO NOT add more than two brackets per single row, for instance never do [John Doe R 3.5x] TURNING INTO [John Doe | 3.5x] | [Robert Brewer | 4.5x].\n DO NOT output any $ signs in a row, for instance [John Doe $1,300,000] should be [John Doe | 1,300,000] \n Count the number of categories and never add extra columns to each row, for instance if their are three categories and the data is [Primitive Bits Inc. Digital Markets / Media 20-Jun $225,000] the data is incorrectly interpretted if it looks like this [Primitive Bits Inc., Digital Markets / Media, 20-Jun, $225,000, - , - ] because it has two - added to the end, making it 5 columns in one row that was supposed to have three columns based off the three categories, the data is correctly interpretted if it looks like this [Primitive Bits Inc., Digital Markets / Media, 20-Jun, $225,000] \n DO NOT EVER ADD ANYTHING OUTSIDE OF THE BRACKETS, for instance this would be wrong [Primitive Bits Inc., Digital Markets / Media, 20-Jun, $225,000],,, \n NEVER HAVE A COMMA OUTSIDE OF A BRACKET LIKE THIS ], \n Always make - into a 0, for instance [Primitive Bits Inc. Digital Markets / Media, 20-Jun - - - - ] TURNS INTO [Primitive Bits Inc. | Digital Markets / Media | 20-Jun | 0 | 0 | 0 | 0 ] \n If an input is - with no adjacent values, assign it to the category that is being searched for; if the categories were name and MOIC, [John Doe - ] TURNS INTO [John Doe | - ]. Do not ever miss a - with no adjacent values. \n If an input is 0 with no adjacent values, assign it to the category that is being searched for; if the categories were name and MOIC, [John Doe 0 ] TURNS INTO [John Doe | 0 ]. Do not ever miss a 0 with no adjacent values. \n The category called category is never over 4 words long. The Name category will never include the words equity, token, or conviertible in it. \n Make certain to fix spelling errors. Make certain to always put each row in brackets.\n NEVER EVER CREATE write in more | than the number of categories minus one, for example if their are 3 categories: name, MOIV, total investment, then the output would be [Robert | 2.6x | 1,200,000].\n Here are additional examples of outputs based on inputs: [John Doe 3.5x] TURNS INTO [John Doe | 3.5x]. \n Another example would be [Mlxar Inc. (Atlas) Gaming/XR 7/8/2022 $ 1,986,500 $ 3,304,499 1.7x SAFE $10.0] TURNS INTO [Mlxar Inc. (Atlas)| Gaming/XR| 7/8/2022| $ 1,986,500| $ 3,304,499| 1.7x| SAFE| $10.0]. \n Here is the data: \n\n"
            
            prompt = '''here is my prompt:

Formulate the data into a list delimited by | based on these categories: {CATEGORIES}. DO NOT CHANGE THE CATEGORY NAMES. Output only a single row. ONLY RETURN each row as a single line. DO NOT USE backslash n. DO NOT WRITE TEXT ABOUT THE OUTPUT.

DO NOT return single standalone letters that do not fit into the respective categories. For example, [John Doe R 3.5x] should turn into [John Doe | 3.5x]
DO NOT add more than two brackets per single row. For example, [John Doe R 3.5x] should turn into [John Doe | 3.5x] and not [John Doe | 3.5x] | [Robert Brewer | 4.5x]
DO NOT output any $ signs in a row. For example, [John Doe $1,300,000] should be [John Doe | 1,300,000]
Count the number of categories and never add extra columns to each row. For instance, if there are three categories and the data is [Primitive Bits Inc. Digital Markets / Media 20-Jun $225,000], it should be interpreted correctly as [Primitive Bits Inc. | Digital Markets / Media | 20-Jun | $225,000] without extra columns.
DO NOT add anything outside of the brackets. For instance, [Primitive Bits Inc., Digital Markets / Media, 20-Jun, $225,000],,, is incorrect.
Always replace - with 0. For example, [Primitive Bits Inc. Digital Markets / Media, 20-Jun - - - -] should turn into [Primitive Bits Inc. | Digital Markets / Media | 20-Jun | 0 | 0 | 0 | 0]
Always replace $0 or 0 with 0. For example, [Primitive Bits Inc. Digital Markets / Media, 20-Jun $0 0 $0 0] should turn into [Primitive Bits Inc. | Digital Markets / Media | 20-Jun | 0 | 0 | 0 | 0]
If an input is 0 with no adjacent values, assign it to the corresponding category. For example, if the categories were name and MOIC, [John Doe 0] should turn into [John Doe | 0]
The category called CATEGORY is never over four words long.
The Company Name or Name category will never include the words equity, token, or convertible.
The Type Of Asset or Structure category will always be equity, token, or convertible. For example ["Vega Protocol Token DeFi The derivatives scaling layer for Web3 $0 - - $292,500 February 2019 $24,000,000 $0 - - $150,000 November 2020 $52,000,000 $0 - - $0 - - 814,752.53 - 171,916.79 172,057.97 1,158,727.29 $216,558 $659,058 - $50,602,500 $902,069 - 1.78%"] would be [Vega Protocol | Token | DeFi | The derivatives scaling layer for Web3 | 0 | 0 | 0 | 292,500 | February 2019 | 24,000,000 | 0 | 0 | 0 | 150,000 | November 2020 | 52,000,000 | 0 | 0 | 0 | 0 | 0 | 0 | 814,752.53 | 0 | 171,916.79 | 172,057.97 | 1,158,727.29 | 216,558 | 659,058 | 0 | 50,602,500 | 902,069 | 0 | 1.78%]
Fix spelling errors.
Always enclose each row in brackets.
Do Not Remove Any Data for instance [Company_1 Tech 05/15/2022 6.03 99.96 2.4 Series D 320.22 91.13 Oceania Token] turns into [Company_1 | Tech | 05/15/2022 | 6.03 | 99.96 | 2.4 | Series D | 320.22 | 91.13 | Oceania | Token] not be [Company_1 | Tech | 05/15/2022] because that removes data. 
Examples:
Categories are "Name,MOIC" and the number of categories is 2, meaning that [John Doe 3.5x] turns into [John Doe | 3.5x].
Categories are "COMPANY NAME, CATEGORY, ENTRY DATE, COST, TOTAL VALUE, GROSS MOIC, ENTRY ROUND, ENTRY VALUATIONS (in millions)" and the number of categories is 8, meaning that[Mlxar Inc. (Atlas) Gaming/XR 7/8/2022 $1,986,500 $3,304,499 1.7x SAFE $10.0] turns into [Mlxar Inc. (Atlas) | Gaming/XR | 7/8/2022 | 1,986,500 | 3,304,499 | 1.7x | SAFE | 10.0]
Categories are and "Name,Type of Asset, Category, Description, Pre-Seed investment ($), Pre-Seed Investment Date, Pre-Seed Valuation (Post), Seed Investment ($), Seed Investment Date, Seed Valuation (post),  Internal Round Investment ($), Internal Round Investment Date, Internal Round Valuation (post), Series A investment ($), Series A Investment Date, Series A Valuation (post), Series A+ investment ($), Series A+ Investment Date, Series A+ Valuation (post), Series B investment ($), Series B Investment Date, Series B Valuation, Vested Tokens, Unvested Tokens, Earned Tokens, Market Purchased Tokens, Total Tokens, OTC/Market Purchase USD, Total Capital Committed ($), Current Equity FDV ($), Current Token FDV ($), Current NAV ($), Equity Ownership (%), Token Ownership (%)*" and the number of categories is 34, meaning that [EXIT Hal Equity Bridge  Blockchain Notification Tool $0 - - "400,000" October 2021 "15,000,000" $0 - -  $0 - - $0 - - - - - - - - - - "400,000" - - "428,000" - -] turns into [EXIT Hal | Equity | Bridge | Blockchain Notification Tool| 0 | 0 | 0 | "400,000" | October 2021 | "15,000,000" | 0 | 0 | 0 | 0 | 0 | 0 | 0 | 0 | 0 | 0 | 0 | 0 | 0 | 0 | 0 | 0 | 0 | 0 | "400,000" | 0 | 0 | "428,000" | 0 | 0]

Each row needs to have {NUMBER_CATEGORIES} items.
Take your time to complete this task, Here are the rows:'''
            for item in csv_reader:
                prompt += f"{item}\n"
                
            # for segment in data:
                # Call the ChatGPT API
            
                response = openai.chat.completions.create(
                    # model="gpt-4o",
                    model="gpt-4-turbo",
                    # model="gpt-3.5-turbo", 
                    
                    messages=[
                        {"role": "system", "content": prompt},
                        {"role": "user", "content": item[0]}
                    ]
                )
                print(f"AI MODEL Working:{i}")
                
                i +=1
                # # Extract the response text
                table = json.loads(response.json())
                print(table["choices"][0]["message"]["content"])
                responses.append(table["choices"][0]["message"]["content"])
            
            return responses
def wait_for_fund_name():
    
    user_input = input("Input the name of tbe fund, hit enter to continue: ")
    return user_input

FUNDNAME = wait_for_fund_name()

def formatting(table):
    print("Formatting Beginning")
    chart = []

    for row in table:
        item = row.split("|")
        for i in range(len(item)):
            item[i] = item[i].strip(" $,[]")
        chart.append(item)

    with open('./na.csv', 'w', newline='') as file:
        writer = csv.writer(file)
        
        # Write the data
        writer.writerow(CATEGORIES.split(","))
        writer.writerows(chart)

    wait_for_yes()

    df = pd.read_csv("na.csv", delimiter=",")
    df.to_excel("data_lake.xlsx", index=False)

    print("Formatting Ending", "\n")
    print("Go To PreFormatted Excel To Enter Data")
    print("Go To data_lake Excel To Copy Data Into PreFormatted")
            

table = format_data_into_table("./na.csv")
formatting(table)



# BIG BOY XLS SHIT


























# Create a new Workbook and select the active Worksheet
wb = Workbook()
ws = wb.active
ws.title = f"{FUNDNAME}"

def Format_Main_Table():
    # Define the data and headers
    headers = ["Index", "Company","Sector","Structure","Date","Invested","Realized","NAV (write current one cell above)","Total Value","MOIC","% Cost","% NAV","Entry Own%","Current Own%", "Lead","Region","Stage","BoD","Implied Entry Val (M)","Implied Current Val (M)","Avg Entry","Year","Tokens","Ticker","   ","Price","<ENTER DATE> NAV","CURRENT NAV","△ NAV","% △ NAV"]

    data = []

    i = 0
    with open('na.csv', mode='r') as file:
            # Create a CSV reader object
            csv_reader = csv.reader(file, delimiter=",")
            
            for row in (csv_reader):
                i = i + 1
                data.append([i])
            for n in range(5):
                data.append(["EMPTY"])
            data.append(["TOTAL"])
            
            

    # Append headers to the worksheet starting at column A, row 4
    header_row = 4
    header_col = 1  # Column A

    for col_num, header in enumerate(headers, start=header_col):
        ws.cell(row=header_row, column=col_num, value=header)

    # Append data to the worksheet starting at column B, row 5
    data_start_row = 5

    for row_num, row_data in enumerate(data, start=data_start_row):
        for col_num, cell_value in enumerate(row_data, start=header_col+1):
            ws.cell(row=row_num, column=col_num-1, value=cell_value)

    
    # Add formula to calculate Total Compensation in each row
    for row_num in range(data_start_row, data_start_row + len(data)-1):
        
        # formatting NAV
        CURRENT_NAV = f"AB{row_num}"
        DATE_NAV = f"AA{row_num}"
        ACTUAL_NAV = f"H{row_num}"
        ws[ACTUAL_NAV] = f"=IF($H$3=\"current\",{DATE_NAV},{CURRENT_NAV})"

        # formatting total value
        REALIZED = f"G{row_num}"
        ACTUAL_NAV = f"H{row_num}"
        TOTAL_VALUE = f"I{row_num}"
        ws[TOTAL_VALUE] = f"={REALIZED}+{ACTUAL_NAV}"

        # formatting MOIC
        INVESTED = f"F{row_num}"
        MOIC = f"J{row_num}"
        ws[MOIC] = f"=IFERROR({TOTAL_VALUE}/{INVESTED},\"n/a\")"

        # formatting % cost
        TOTAL_INVESTED = f"$F${data_start_row+len(data)-1}"
        PERCENT_COST = f"K{row_num}"
        ws[PERCENT_COST] = f"={INVESTED}/{TOTAL_INVESTED}"

        # formatting % NAV
        TOTAL_NAV = f"$H${data_start_row+len(data)-1}"
        PERCENT_NAV = f"L{row_num}"
        ws[PERCENT_NAV] = f"={ACTUAL_NAV}/{TOTAL_NAV}"

        # formatting Implied Entry Valuation % (M)
        ENTRY_OWN_PERCENT = f"M{row_num}"
        IMPLIED_ENTRY = f"S{row_num}"
        ws[IMPLIED_ENTRY] = f"=IFERROR({INVESTED}/{ENTRY_OWN_PERCENT}/1000000,\"n/a\")"

        # formatting Current Entry Valuation % (M)
        CURRENT_OWN_PERCENT = f"N{row_num}"
        IMPLIED_CURRENT_VALUE = f"T{row_num}"
        ws[IMPLIED_CURRENT_VALUE] = f"=IFERROR({ACTUAL_NAV}/{CURRENT_OWN_PERCENT}/1000000,\"n/a\")"

        # formatting year of investment
        DATE = f"E{row_num}"
        INV_YEAR = f"V{row_num}"
        ws[INV_YEAR] = f"=YEAR({DATE})"

        # formatting change in NAV
        CHANGE_NAV = f"AC{row_num}"
        ws[CHANGE_NAV] = f"={DATE_NAV}-{CURRENT_NAV}"

        # formatting percent change in NAV between DATE and CURRENT
        PERCENTAGE_DATE_CURRENT = f"AD{row_num}"
        ws[PERCENTAGE_DATE_CURRENT] = f"={CHANGE_NAV}/{DATE_NAV}-1"

        # formatting Avg Entry
        AVG_ENTRY = f"U{row_num}"
        ws[AVG_ENTRY] = f"=IFERROR({IMPLIED_CURRENT_VALUE}*{INVESTED}/{TOTAL_VALUE},\"n/a\")"

    #TOTALS ROW
    # formatting invested total
    INVESTING_TOTAL = f"F{data_start_row+len(data)-1}"
    INVESTING_TOTAL_ROWS = f"F{data_start_row}:F{data_start_row+len(data)-2}"
    ws[INVESTING_TOTAL] = f"=SUM({INVESTING_TOTAL_ROWS})"

    # formatting realized total
    REALIZED_TOTAL = f"G{data_start_row+len(data)-1}"
    REALIZED_TOTAL_ROWS = f"G{data_start_row}:G{data_start_row+len(data)-2}"
    ws[REALIZED_TOTAL] = f"=SUM({REALIZED_TOTAL_ROWS})"

    # formatting NAV total
    NAV_TOTAL = f"H{data_start_row+len(data)-1}"
    NAV_TOTAL_ROWS = f"H{data_start_row}:H{data_start_row+len(data)-2}"
    ws[NAV_TOTAL] = f"=SUM({NAV_TOTAL_ROWS})"

    # formatting SUM TOTAL VALUE 
    TOTAL_VALUE_TOTAL = f"I{data_start_row+len(data)-1}"
    TOTAL_VALUE_TOTAL_ROWS = f"I{data_start_row}:I{data_start_row+len(data)-2}"
    ws[TOTAL_VALUE_TOTAL] = f"=SUM({TOTAL_VALUE_TOTAL_ROWS})"

    # formatting MOIC
    MOIC_ = f"J{data_start_row+len(data)-1}"
    ws[MOIC_] = f"=IFERROR({TOTAL_VALUE_TOTAL}/{INVESTING_TOTAL},\"n/a\")"

    # formatting % cost
    PERCENT_COST_ = f"K{data_start_row+len(data)-1}"
    ws[PERCENT_COST_] = f"={INVESTING_TOTAL}/{INVESTING_TOTAL}"

    # formatting % NAV
    PERCENT_NAV_ = f"L{data_start_row+len(data)-1}"
    ws[PERCENT_NAV_] = f"={NAV_TOTAL}/{NAV_TOTAL}"

    # misc equation
    IMPLIED_ENTRY_ROWS = f"S{data_start_row}:S{data_start_row+len(data)-2}"
    MISC_EQ = f"S{data_start_row+len(data)-1}"
    ws[MISC_EQ] = f"=SUMPRODUCT({INVESTING_TOTAL_ROWS},{IMPLIED_ENTRY_ROWS})/{INVESTING_TOTAL}"

    return data_start_row+len(data)-2


GLOBAL_COUNT = Format_Main_Table()

WB2=f"{FUNDNAME} Pivot Tables"
ws2 = wb.create_sheet(WB2)
wb.active = ws2

GLOBAL_DATA = [
    [1],
    [2],
    [3],
    [4],
    [5],
    [6],
    [7],
    [8],
    [9],
    [10],
    ["TOTAL"]
]




INVESTED = f"\'{FUNDNAME}\'!$F$5:$F${GLOBAL_COUNT}"
REALIZED = f"\'{FUNDNAME}\'!$G$5:$G${GLOBAL_COUNT}"
NAV = f"\'{FUNDNAME}\'!$H$5:$H${GLOBAL_COUNT}"
TOTAL_VALUE = f"\'{FUNDNAME}\'!$I$5:$I${GLOBAL_COUNT}"


def year():
    headers = ["Index","Year", "Count", "Invested", "Realized", "NAV", "Total Value", "MOIC", "% Cost", "% NAV"]
     
    # Append headers to the worksheet starting at column A, row 4
    header_row = 3
    header_col = 1  # Column A

    for col_num, header in enumerate(headers, start=header_col):
        ws2.cell(row=header_row, column=col_num, value=header)

    # Append data to the worksheet starting at column B, row 5
    data_start_row = 4

    # CHANGE
    YEAR = f"\'{FUNDNAME}\'!$V$5:$V${GLOBAL_COUNT}"

    for row_num, row_data in enumerate(GLOBAL_DATA, start=data_start_row):
        for col_num, cell_value in enumerate(row_data, start=header_col+1):
            ws2.cell(row=row_num, column=col_num-1, value=cell_value)

    for row_num in range(data_start_row, data_start_row + len(GLOBAL_DATA)-1):
        # count if year
        CAT_YEAR = f"B{row_num}"
        COUNT = f"C{row_num}"
        ws2[COUNT] = f"=COUNTIF({YEAR},\'{WB2}\'!{CAT_YEAR})"

        # sum if year by investment
        INVESTED_ = f"D{row_num}"
        ws2[INVESTED_] = f"=SUMIF({YEAR},\'{WB2}\'!{CAT_YEAR},{INVESTED})"

        # sum if year by realized
        REALIZED_ = f"E{row_num}"
        ws2[REALIZED_] = f"=SUMIF({YEAR},\'{WB2}\'!{CAT_YEAR},{REALIZED})"       

        # sum if year by NAV
        NAV_ = f"F{row_num}"
        ws2[NAV_] = f"=SUMIF({YEAR},\'{WB2}\'!{CAT_YEAR},{NAV})"   

        # sum if year by NAV
        TOTAL_VALUE_ = f"G{row_num}"
        ws2[TOTAL_VALUE_] = f"=SUMIF({YEAR},\'{WB2}\'!{CAT_YEAR},{TOTAL_VALUE})"   

        # MOIC calculation
        MOIC = f"H{row_num}"
        ws2[MOIC] = f"={TOTAL_VALUE_}/{INVESTED_}"

        # %Cost
        TOTAL_INVESTED = f"$D${data_start_row+len(GLOBAL_DATA)-1}"
        PERCENT_COST = f"I{row_num}"
        ws2[PERCENT_COST] = f"={INVESTED_}/{TOTAL_INVESTED}"

        # %NAV
        TOTAL_NAV = f"$F${data_start_row+len(GLOBAL_DATA)-1}"
        PERCENT_NAV = f"J{row_num}"
        ws2[PERCENT_NAV] = f"={NAV_}/{TOTAL_NAV}"

    # counting
    COUNTING_TOTAL = f"C{data_start_row+len(GLOBAL_DATA)-1}"
    COUNTING_SUM_ROWS = f"C{data_start_row}:C{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[COUNTING_TOTAL] = f"=SUM({COUNTING_SUM_ROWS})"
    
    # investing total rows
    INVESTING_TOTAL = f"D{data_start_row+len(GLOBAL_DATA)-1}"
    INVESTING_TOTAL_ROWS = f"D{data_start_row}:D{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[INVESTING_TOTAL] = f"=SUM({INVESTING_TOTAL_ROWS})"
     
    # formatting realized total
    REALIZED_TOTAL = f"E{data_start_row+len(GLOBAL_DATA)-1}"
    REALIZED_TOTAL_ROWS = f"E{data_start_row}:E{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[REALIZED_TOTAL] = f"=SUM({REALIZED_TOTAL_ROWS})"

    # formatting NAV total
    NAV_TOTAL = f"F{data_start_row+len(GLOBAL_DATA)-1}"
    NAV_TOTAL_ROWS = f"F{data_start_row}:F{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[NAV_TOTAL] = f"=SUM({NAV_TOTAL_ROWS})"

    # formatting SUM TOTAL VALUE 
    TOTAL_VALUE_TOTAL = f"G{data_start_row+len(GLOBAL_DATA)-1}"
    TOTAL_VALUE_TOTAL_ROWS = f"G{data_start_row}:G{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[TOTAL_VALUE_TOTAL] = f"=SUM({TOTAL_VALUE_TOTAL_ROWS})"

    # formatting MOIC
    MOIC_ = f"H{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[MOIC_] = f"=IFERROR({TOTAL_VALUE_TOTAL}/{INVESTING_TOTAL},\"n/a\")"

    # formatting % cost
    PERCENT_COST_ = f"I{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[PERCENT_COST_] = f"={INVESTING_TOTAL}/{INVESTING_TOTAL}"

    # formatting % NAV
    PERCENT_NAV_ = f"J{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[PERCENT_NAV_] = f"={NAV_TOTAL}/{NAV_TOTAL}"





def Stage():
    
    headers = ["Index","Stage", "Count", "Invested", "Realized", "NAV", "Total Value", "MOIC", "% Cost", "% NAV"]
    
    # Append headers to the worksheet starting at column A, row 4
    header_row = 17
    header_col = 1  # Column A

    for col_num, header in enumerate(headers, start=header_col):
        ws2.cell(row=header_row, column=col_num, value=header)

    # Append data to the worksheet starting at column B, row 5
    data_start_row = 18

    # CHANGE
    CAT = f"\'{FUNDNAME}\'!$Q$5:$Q${GLOBAL_COUNT}"

    for row_num, row_data in enumerate(GLOBAL_DATA, start=data_start_row):
        for col_num, cell_value in enumerate(row_data, start=header_col+1):
            ws2.cell(row=row_num, column=col_num-1, value=cell_value)

    for row_num in range(data_start_row, data_start_row + len(GLOBAL_DATA)-1):
        # count if year
        CAT_STAGE = f"B{row_num}"
        COUNT = f"C{row_num}"
        ws2[COUNT] = f"=COUNTIF({CAT},\'{WB2}\'!{CAT_STAGE})"

        # sum if year by investment
        INVESTED_ = f"D{row_num}"
        ws2[INVESTED_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{INVESTED})"

        # sum if year by realized
        REALIZED_ = f"E{row_num}"
        ws2[REALIZED_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{REALIZED})"       

        # sum if year by NAV
        NAV_ = f"F{row_num}"
        ws2[NAV_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{NAV})"   

        # sum if year by NAV
        TOTAL_VALUE_ = f"G{row_num}"
        ws2[TOTAL_VALUE_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{TOTAL_VALUE})"   

        # MOIC calculation
        MOIC = f"H{row_num}"
        ws2[MOIC] = f"={TOTAL_VALUE_}/{INVESTED_}"

        # %Cost
        TOTAL_INVESTED = f"$D${data_start_row+len(GLOBAL_DATA)-1}"
        PERCENT_COST = f"I{row_num}"
        ws2[PERCENT_COST] = f"={INVESTED_}/{TOTAL_INVESTED}"

        # %NAV
        TOTAL_NAV = f"$F${data_start_row+len(GLOBAL_DATA)-1}"
        PERCENT_NAV = f"J{row_num}"
        ws2[PERCENT_NAV] = f"={NAV_}/{TOTAL_NAV}"
    
    # counting
    COUNTING_TOTAL = f"C{data_start_row+len(GLOBAL_DATA)-1}"
    COUNTING_SUM_ROWS = f"C{data_start_row}:C{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[COUNTING_TOTAL] = f"=SUM({COUNTING_SUM_ROWS})"

    # investing total rows
    INVESTING_TOTAL = f"D{data_start_row+len(GLOBAL_DATA)-1}"
    INVESTING_TOTAL_ROWS = f"D{data_start_row}:D{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[INVESTING_TOTAL] = f"=SUM({INVESTING_TOTAL_ROWS})"
     
    # formatting realized total
    REALIZED_TOTAL = f"E{data_start_row+len(GLOBAL_DATA)-1}"
    REALIZED_TOTAL_ROWS = f"E{data_start_row}:E{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[REALIZED_TOTAL] = f"=SUM({REALIZED_TOTAL_ROWS})"

    # formatting NAV total
    NAV_TOTAL = f"F{data_start_row+len(GLOBAL_DATA)-1}"
    NAV_TOTAL_ROWS = f"F{data_start_row}:F{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[NAV_TOTAL] = f"=SUM({NAV_TOTAL_ROWS})"

    # formatting SUM TOTAL VALUE 
    TOTAL_VALUE_TOTAL = f"G{data_start_row+len(GLOBAL_DATA)-1}"
    TOTAL_VALUE_TOTAL_ROWS = f"G{data_start_row}:G{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[TOTAL_VALUE_TOTAL] = f"=SUM({TOTAL_VALUE_TOTAL_ROWS})"

    # formatting MOIC
    MOIC_ = f"H{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[MOIC_] = f"=IFERROR({TOTAL_VALUE_TOTAL}/{INVESTING_TOTAL},\"n/a\")"

    # formatting % cost
    PERCENT_COST_ = f"I{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[PERCENT_COST_] = f"={INVESTING_TOTAL}/{INVESTING_TOTAL}"

    # formatting % NAV
    PERCENT_NAV_ = f"J{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[PERCENT_NAV_] = f"={NAV_TOTAL}/{NAV_TOTAL}"




def Geography():
    
    headers = ["Index","Geography", "Count", "Invested", "Realized", "NAV", "Total Value", "MOIC", "% Cost", "% NAV"]
    
    # Append headers to the worksheet starting at column A, row 4
    header_row = 31
    header_col = 1  # Column A

    for col_num, header in enumerate(headers, start=header_col):
        ws2.cell(row=header_row, column=col_num, value=header)

    # Append data to the worksheet starting at column B, row 5
    data_start_row = 32

    # CHANGE
    CAT = f"\'{FUNDNAME}\'!$P$5:$P${GLOBAL_COUNT}"

    for row_num, row_data in enumerate(GLOBAL_DATA, start=data_start_row):
        for col_num, cell_value in enumerate(row_data, start=header_col+1):
            ws2.cell(row=row_num, column=col_num-1, value=cell_value)

    for row_num in range(data_start_row, data_start_row + len(GLOBAL_DATA)-1):
        # count if year
        CAT_STAGE = f"B{row_num}"
        COUNT = f"C{row_num}"
        ws2[COUNT] = f"=COUNTIF({CAT},\'{WB2}\'!{CAT_STAGE})"

        # sum if year by investment
        INVESTED_ = f"D{row_num}"
        ws2[INVESTED_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{INVESTED})"

        # sum if year by realized
        REALIZED_ = f"E{row_num}"
        ws2[REALIZED_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{REALIZED})"       

        # sum if year by NAV
        NAV_ = f"F{row_num}"
        ws2[NAV_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{NAV})"   

        # sum if year by NAV
        TOTAL_VALUE_ = f"G{row_num}"
        ws2[TOTAL_VALUE_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{TOTAL_VALUE})"   

        # MOIC calculation
        MOIC = f"H{row_num}"
        ws2[MOIC] = f"={TOTAL_VALUE_}/{INVESTED_}"


        # %Cost
        TOTAL_INVESTED = f"$D${data_start_row+len(GLOBAL_DATA)-1}"
        PERCENT_COST = f"I{row_num}"
        ws2[PERCENT_COST] = f"={INVESTED_}/{TOTAL_INVESTED}"

        # %NAV
        TOTAL_NAV = f"$F${data_start_row+len(GLOBAL_DATA)-1}"
        PERCENT_NAV = f"J{row_num}"
        ws2[PERCENT_NAV] = f"={NAV_}/{TOTAL_NAV}"
    
    # counting
    COUNTING_TOTAL = f"C{data_start_row+len(GLOBAL_DATA)-1}"
    COUNTING_SUM_ROWS = f"C{data_start_row}:C{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[COUNTING_TOTAL] = f"=SUM({COUNTING_SUM_ROWS})"

    # investing total rows
    INVESTING_TOTAL = f"D{data_start_row+len(GLOBAL_DATA)-1}"
    INVESTING_TOTAL_ROWS = f"D{data_start_row}:D{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[INVESTING_TOTAL] = f"=SUM({INVESTING_TOTAL_ROWS})"
     
    # formatting realized total
    REALIZED_TOTAL = f"E{data_start_row+len(GLOBAL_DATA)-1}"
    REALIZED_TOTAL_ROWS = f"E{data_start_row}:E{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[REALIZED_TOTAL] = f"=SUM({REALIZED_TOTAL_ROWS})"

    # formatting NAV total
    NAV_TOTAL = f"F{data_start_row+len(GLOBAL_DATA)-1}"
    NAV_TOTAL_ROWS = f"F{data_start_row}:F{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[NAV_TOTAL] = f"=SUM({NAV_TOTAL_ROWS})"

    # formatting SUM TOTAL VALUE 
    TOTAL_VALUE_TOTAL = f"G{data_start_row+len(GLOBAL_DATA)-1}"
    TOTAL_VALUE_TOTAL_ROWS = f"G{data_start_row}:G{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[TOTAL_VALUE_TOTAL] = f"=SUM({TOTAL_VALUE_TOTAL_ROWS})"

    # formatting MOIC
    MOIC_ = f"H{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[MOIC_] = f"=IFERROR({TOTAL_VALUE_TOTAL}/{INVESTING_TOTAL},\"n/a\")"

    # formatting % cost
    PERCENT_COST_ = f"I{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[PERCENT_COST_] = f"={INVESTING_TOTAL}/{INVESTING_TOTAL}"

    # formatting % NAV
    PERCENT_NAV_ = f"J{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[PERCENT_NAV_] = f"={NAV_TOTAL}/{NAV_TOTAL}"



def Structure():
    
    headers = ["Index","Structure", "Count", "Invested", "Realized", "NAV", "Total Value", "MOIC", "% Cost", "% NAV"]
    
    # Append headers to the worksheet starting at column A, row 4
    header_row = 45
    header_col = 1  # Column A

    for col_num, header in enumerate(headers, start=header_col):
        ws2.cell(row=header_row, column=col_num, value=header)

    # Append data to the worksheet starting at column B, row 5
    data_start_row = 46

    # CHANGE
    CAT = f"\'{FUNDNAME}\'!$D$5:$D${GLOBAL_COUNT}"

    for row_num, row_data in enumerate(GLOBAL_DATA, start=data_start_row):
        for col_num, cell_value in enumerate(row_data, start=header_col+1):
            ws2.cell(row=row_num, column=col_num-1, value=cell_value)

    for row_num in range(data_start_row, data_start_row + len(GLOBAL_DATA)-1):
        # count if year
        CAT_STAGE = f"B{row_num}"
        COUNT = f"C{row_num}"
        ws2[COUNT] = f"=COUNTIF({CAT},\'{WB2}\'!{CAT_STAGE})"

        # sum if year by investment
        INVESTED_ = f"D{row_num}"
        ws2[INVESTED_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{INVESTED})"

        # sum if year by realized
        REALIZED_ = f"E{row_num}"
        ws2[REALIZED_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{REALIZED})"       

        # sum if year by NAV
        NAV_ = f"F{row_num}"
        ws2[NAV_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{NAV})"   

        # sum if year by NAV
        TOTAL_VALUE_ = f"G{row_num}"
        ws2[TOTAL_VALUE_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{TOTAL_VALUE})"   

        # MOIC calculation
        MOIC = f"H{row_num}"
        ws2[MOIC] = f"={TOTAL_VALUE_}/{INVESTED_}"


        # %Cost
        TOTAL_INVESTED = f"$D${data_start_row+len(GLOBAL_DATA)-1}"
        PERCENT_COST = f"I{row_num}"
        ws2[PERCENT_COST] = f"={INVESTED_}/{TOTAL_INVESTED}"

        # %NAV
        TOTAL_NAV = f"$F${data_start_row+len(GLOBAL_DATA)-1}"
        PERCENT_NAV = f"J{row_num}"
        ws2[PERCENT_NAV] = f"={NAV_}/{TOTAL_NAV}"
    
    # counting
    COUNTING_TOTAL = f"C{data_start_row+len(GLOBAL_DATA)-1}"
    COUNTING_SUM_ROWS = f"C{data_start_row}:C{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[COUNTING_TOTAL] = f"=SUM({COUNTING_SUM_ROWS})"

    # investing total rows
    INVESTING_TOTAL = f"D{data_start_row+len(GLOBAL_DATA)-1}"
    INVESTING_TOTAL_ROWS = f"D{data_start_row}:D{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[INVESTING_TOTAL] = f"=SUM({INVESTING_TOTAL_ROWS})"
     
    # formatting realized total
    REALIZED_TOTAL = f"E{data_start_row+len(GLOBAL_DATA)-1}"
    REALIZED_TOTAL_ROWS = f"E{data_start_row}:E{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[REALIZED_TOTAL] = f"=SUM({REALIZED_TOTAL_ROWS})"

    # formatting NAV total
    NAV_TOTAL = f"F{data_start_row+len(GLOBAL_DATA)-1}"
    NAV_TOTAL_ROWS = f"F{data_start_row}:F{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[NAV_TOTAL] = f"=SUM({NAV_TOTAL_ROWS})"

    # formatting SUM TOTAL VALUE 
    TOTAL_VALUE_TOTAL = f"G{data_start_row+len(GLOBAL_DATA)-1}"
    TOTAL_VALUE_TOTAL_ROWS = f"G{data_start_row}:G{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[TOTAL_VALUE_TOTAL] = f"=SUM({TOTAL_VALUE_TOTAL_ROWS})"

    # formatting MOIC
    MOIC_ = f"H{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[MOIC_] = f"=IFERROR({TOTAL_VALUE_TOTAL}/{INVESTING_TOTAL},\"n/a\")"

    # formatting % cost
    PERCENT_COST_ = f"I{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[PERCENT_COST_] = f"={INVESTING_TOTAL}/{INVESTING_TOTAL}"

    # formatting % NAV
    PERCENT_NAV_ = f"J{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[PERCENT_NAV_] = f"={NAV_TOTAL}/{NAV_TOTAL}"




def EntryOwenership():
    
    headers = ["Index","Entry Owenership", "Count", "Invested", "Realized", "NAV", "Total Value", "MOIC", "% Cost", "% NAV"]
    
    # Append headers to the worksheet starting at column A, row 4
    header_row = 59
    header_col = 1  # Column A

    for col_num, header in enumerate(headers, start=header_col):
        ws2.cell(row=header_row, column=col_num, value=header)

    # Append data to the worksheet starting at column B, row 5
    data_start_row = 60

    # CHANGE
    CAT = f"\'{FUNDNAME}\'!$M$5:$M${GLOBAL_COUNT}"

    for row_num, row_data in enumerate(GLOBAL_DATA, start=data_start_row):
        for col_num, cell_value in enumerate(row_data, start=header_col+1):
            ws2.cell(row=row_num, column=col_num-1, value=cell_value)



    for row_num in range(data_start_row, data_start_row + len(GLOBAL_DATA)-1):
        # count if year
        COUNT = f"C{row_num}"

        if row_num == 60:
            ws2["B60"]=" x <= 1%"
            ws2[COUNT] = f"=COUNTIF({CAT},\"<=1%\")"
            # sum if year by investment
            INVESTED_ = f"D{row_num}"
            ws2[INVESTED_] = f"=SUMIF({CAT},\"<=1%\",{INVESTED})"

            # sum if year by realized
            REALIZED_ = f"E{row_num}"
            ws2[REALIZED_] = f"=SUMIF({CAT},\"<=1%\",{REALIZED})"       

            # sum if year by NAV
            NAV_ = f"F{row_num}"
            ws2[NAV_] = f"=SUMIF({CAT},\"<=1%\",{NAV})"   

            # sum if year by NAV
            TOTAL_VALUE_ = f"G{row_num}"
            ws2[TOTAL_VALUE_] = f"=SUMIF({CAT},\"<=1%\",{TOTAL_VALUE})"      
        elif row_num == 61:
            ws2["B61"]="1% < x <= 5%"
            ws2[COUNT] = f"=COUNTIF({CAT},\"<=5%\")-COUNTIF({CAT},\"<=1%\")"
            # sum if year by investment
            SUB_INVESTED_ = f"D{row_num}"
            INVESTED_ = f"D{row_num}"
            ws2[INVESTED_] = f"=SUMIF({CAT},\"<=5%\",{INVESTED})-{SUB_INVESTED_}"

            # sum if year by realized
            SUB_REALIZED_ = f"E{row_num-1}"
            REALIZED_ = f"E{row_num}"
            ws2[REALIZED_] = f"=SUMIF({CAT},\"<=5%\",{REALIZED})-{SUB_REALIZED_}"       

            # sum if year by NAV
            SUB_NAV_ = f"F{row_num-1}"
            NAV_ = f"F{row_num}"
            ws2[NAV_] = f"=SUMIF({CAT},\"<=5%\",{NAV})-{SUB_NAV_}"   

            # sum if year by NAV
            SUB_TOTAL_VALUE_ = f"G{row_num-1}"
            TOTAL_VALUE_ = f"G{row_num}"
            ws2[TOTAL_VALUE_] = f"=SUMIF({CAT},\"<=5%\",{TOTAL_VALUE})-{SUB_TOTAL_VALUE_}"               
        elif row_num == 62: 
            ws2["B62"]="x >= 5%"

            ws2[COUNT] = f"=COUNTIF({CAT},\">5%\")"
            
            ws2[COUNT] = f"=COUNTIF({CAT},\">5%\")"
            # sum if year by investment
            INVESTED_ = f"D{row_num}"
            ws2[INVESTED_] = f"=SUMIF({CAT},\">5%\",{INVESTED})"

            # sum if year by realized
            REALIZED_ = f"E{row_num}"
            ws2[REALIZED_] = f"=SUMIF({CAT},\">5%\",{REALIZED})"       

            # sum if year by NAV
            NAV_ = f"F{row_num}"
            ws2[NAV_] = f"=SUMIF({CAT},\">5%\",{NAV})"   

            # sum if year by NAV
            TOTAL_VALUE_ = f"G{row_num}"
            ws2[TOTAL_VALUE_] = f"=SUMIF({CAT},\"<=5%\",{TOTAL_VALUE})"     
        else:
            ws2[COUNT] = f"=COUNTIF({CAT},\"n/a\")" 

            INVESTED_ = f"D{row_num}"
            ws2[INVESTED_] = f"=SUMIF({CAT},\"n/a\",{INVESTED})"

            # sum if year by realized
            REALIZED_ = f"E{row_num}"
            ws2[REALIZED_] = f"=SUMIF({CAT},\"n/a\",{REALIZED})"       

            # sum if year by NAV
            NAV_ = f"F{row_num}"
            ws2[NAV_] = f"=SUMIF({CAT},\"n/a\",{NAV})"   

            # sum if year by NAV
            TOTAL_VALUE_ = f"G{row_num}"
            ws2[TOTAL_VALUE_] = f"=SUMIF({CAT},\"n/a\",{TOTAL_VALUE})"

        # MOIC calculation
        MOIC = f"H{row_num}"
        ws2[MOIC] = f"={TOTAL_VALUE_}/{INVESTED_}"

        # %Cost
        TOTAL_INVESTED = f"$D${data_start_row+len(GLOBAL_DATA)-1}"
        PERCENT_COST = f"I{row_num}"
        ws2[PERCENT_COST] = f"={INVESTED_}/{TOTAL_INVESTED}"

        # %NAV
        TOTAL_NAV = f"$F${data_start_row+len(GLOBAL_DATA)-1}"
        PERCENT_NAV = f"J{row_num}"
        ws2[PERCENT_NAV] = f"={NAV_}/{TOTAL_NAV}"
    
    # counting
    COUNTING_TOTAL = f"C{data_start_row+len(GLOBAL_DATA)-1}"
    COUNTING_SUM_ROWS = f"C{data_start_row}:C{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[COUNTING_TOTAL] = f"=SUM({COUNTING_SUM_ROWS})"

    # investing total rows
    INVESTING_TOTAL = f"D{data_start_row+len(GLOBAL_DATA)-1}"
    INVESTING_TOTAL_ROWS = f"D{data_start_row}:D{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[INVESTING_TOTAL] = f"=SUM({INVESTING_TOTAL_ROWS})"
     
    # formatting realized total
    REALIZED_TOTAL = f"E{data_start_row+len(GLOBAL_DATA)-1}"
    REALIZED_TOTAL_ROWS = f"E{data_start_row}:E{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[REALIZED_TOTAL] = f"=SUM({REALIZED_TOTAL_ROWS})"

    # formatting NAV total
    NAV_TOTAL = f"F{data_start_row+len(GLOBAL_DATA)-1}"
    NAV_TOTAL_ROWS = f"F{data_start_row}:F{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[NAV_TOTAL] = f"=SUM({NAV_TOTAL_ROWS})"

    # formatting SUM TOTAL VALUE 
    TOTAL_VALUE_TOTAL = f"G{data_start_row+len(GLOBAL_DATA)-1}"
    TOTAL_VALUE_TOTAL_ROWS = f"G{data_start_row}:G{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[TOTAL_VALUE_TOTAL] = f"=SUM({TOTAL_VALUE_TOTAL_ROWS})"

    # formatting MOIC
    MOIC_ = f"H{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[MOIC_] = f"=IFERROR({TOTAL_VALUE_TOTAL}/{INVESTING_TOTAL},\"n/a\")"

    # formatting % cost
    PERCENT_COST_ = f"I{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[PERCENT_COST_] = f"={INVESTING_TOTAL}/{INVESTING_TOTAL}"

    # formatting % NAV
    PERCENT_NAV_ = f"J{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[PERCENT_NAV_] = f"={NAV_TOTAL}/{NAV_TOTAL}"



def Sector():
    
    headers = ["Index","Sector", "Count", "Invested", "Realized", "NAV", "Total Value", "MOIC", "% Cost", "% NAV"]
    
    # Append headers to the worksheet starting at column A, row 4
    header_row = 73
    header_col = 1  # Column A

    for col_num, header in enumerate(headers, start=header_col):
        ws2.cell(row=header_row, column=col_num, value=header)

    # Append data to the worksheet starting at column B, row 5
    data_start_row = 74

    # CHANGE
    CAT = f"\'{FUNDNAME}\'!$C$5:$C${GLOBAL_COUNT}"

    for row_num, row_data in enumerate(GLOBAL_DATA, start=data_start_row):
        for col_num, cell_value in enumerate(row_data, start=header_col+1):
            ws2.cell(row=row_num, column=col_num-1, value=cell_value)

    for row_num in range(data_start_row, data_start_row + len(GLOBAL_DATA)-1):
        # count if year
        CAT_STAGE = f"B{row_num}"
        COUNT = f"C{row_num}"
        ws2[COUNT] = f"=COUNTIF({CAT},\'{WB2}\'!{CAT_STAGE})"

        # sum if year by investment
        INVESTED_ = f"D{row_num}"
        ws2[INVESTED_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{INVESTED})"

        # sum if year by realized
        REALIZED_ = f"E{row_num}"
        ws2[REALIZED_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{REALIZED})"       

        # sum if year by NAV
        NAV_ = f"F{row_num}"
        ws2[NAV_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{NAV})"   

        # sum if year by NAV
        TOTAL_VALUE_ = f"G{row_num}"
        ws2[TOTAL_VALUE_] = f"=SUMIF({CAT},\'{WB2}\'!{CAT_STAGE},{TOTAL_VALUE})"   

        # MOIC calculation
        MOIC = f"H{row_num}"
        ws2[MOIC] = f"={TOTAL_VALUE_}/{INVESTED_}"


        # %Cost
        TOTAL_INVESTED = f"$D${data_start_row+len(GLOBAL_DATA)-1}"
        PERCENT_COST = f"I{row_num}"
        ws2[PERCENT_COST] = f"={INVESTED_}/{TOTAL_INVESTED}"

        # %NAV
        TOTAL_NAV = f"$F${data_start_row+len(GLOBAL_DATA)-1}"
        PERCENT_NAV = f"J{row_num}"
        ws2[PERCENT_NAV] = f"={NAV_}/{TOTAL_NAV}"
    
    # counting
    COUNTING_TOTAL = f"C{data_start_row+len(GLOBAL_DATA)-1}"
    COUNTING_SUM_ROWS = f"C{data_start_row}:C{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[COUNTING_TOTAL] = f"=SUM({COUNTING_SUM_ROWS})"

    # investing total rows
    INVESTING_TOTAL = f"D{data_start_row+len(GLOBAL_DATA)-1}"
    INVESTING_TOTAL_ROWS = f"D{data_start_row}:D{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[INVESTING_TOTAL] = f"=SUM({INVESTING_TOTAL_ROWS})"
     
    # formatting realized total
    REALIZED_TOTAL = f"E{data_start_row+len(GLOBAL_DATA)-1}"
    REALIZED_TOTAL_ROWS = f"E{data_start_row}:E{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[REALIZED_TOTAL] = f"=SUM({REALIZED_TOTAL_ROWS})"

    # formatting NAV total
    NAV_TOTAL = f"F{data_start_row+len(GLOBAL_DATA)-1}"
    NAV_TOTAL_ROWS = f"F{data_start_row}:F{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[NAV_TOTAL] = f"=SUM({NAV_TOTAL_ROWS})"

    # formatting SUM TOTAL VALUE 
    TOTAL_VALUE_TOTAL = f"G{data_start_row+len(GLOBAL_DATA)-1}"
    TOTAL_VALUE_TOTAL_ROWS = f"G{data_start_row}:G{data_start_row+len(GLOBAL_DATA)-2}"
    ws2[TOTAL_VALUE_TOTAL] = f"=SUM({TOTAL_VALUE_TOTAL_ROWS})"

    # formatting MOIC
    MOIC_ = f"H{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[MOIC_] = f"=IFERROR({TOTAL_VALUE_TOTAL}/{INVESTING_TOTAL},\"n/a\")"

    # formatting % cost
    PERCENT_COST_ = f"I{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[PERCENT_COST_] = f"={INVESTING_TOTAL}/{INVESTING_TOTAL}"

    # formatting % NAV
    PERCENT_NAV_ = f"J{data_start_row+len(GLOBAL_DATA)-1}"
    ws2[PERCENT_NAV_] = f"={NAV_TOTAL}/{NAV_TOTAL}"



EntryOwenership()
year()
Stage()
Geography()
Structure()
Sector()

# Save the workbook
wb.save("./PreFormattedExcelTemplate.xlsx")
print("the file is saved to PreFormattedExcelTemplate.xlsx in your current folder")